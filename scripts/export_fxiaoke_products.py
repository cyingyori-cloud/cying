#!/usr/bin/env python3
"""
Export Fxiaoke ProductObj records into a JSON file consumable by the prototype.

This script reuses the local Codex skill `fxiaoke-crm-full-readonly-query`
instead of embedding OpenAPI credentials into the frontend project.
"""

from __future__ import annotations

import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SKILL_ROOT = Path("/Users/cying/.codex/skills/fxiaoke-crm-full-readonly-query")
TOOL = SKILL_ROOT / "scripts" / "fxiaoke_openapi_tool.py"
CONFIG = SKILL_ROOT / "config" / "credentials.local.json"
OUTPUT = ROOT / "public" / "data" / "fxiaoke-products.json"
MAPPING_CANDIDATES = [
    Path("/Users/cying/Desktop/crm -- powerquote映射表.xlsx"),
    Path("/Users/cying/Desktop/纷享- powerquote映射表.xlsx"),
]

FIELDS = ",".join([
    "_id",
    "name",
    "model",
    "product_code",
    "field_93IC1__c",
    "category",
    "product_line",
    "product_spec",
    "price",
    "is_saleable",
    "product_status",
    "life_status",
    "on_shelves_time",
    "field_2wOsi__c",
    "product_name_english__c",
    "field_uo1QB__c",
    "description",
    "picture_path",
])

ENERGY_HIGH_KEYWORDS = (
    "储能",
    "电池",
    "pack",
    "ups",
    "高压直流",
)

ENERGY_MEDIUM_KEYWORDS = (
    "电表",
    "ami",
    "数据中心",
    "备电",
    "智能配电",
)

MANUAL_CATEGORY_LABELS = {
    "2": "高压电机",
    "5": "继电保护",
    "6": "交流充电桩",
    "7": "储能变流器",
    "8": "电池类",
    "12": "智能电表",
}


def run_tool(*args: str) -> dict[str, Any]:
    cmd = [
        sys.executable,
        str(TOOL),
        *args,
        "--config",
        str(CONFIG),
    ]
    completed = subprocess.run(
        cmd,
        check=True,
        capture_output=True,
        text=True,
    )
    stdout = completed.stdout
    json_start = stdout.find("{")
    if json_start < 0:
        raise RuntimeError(f"Tool output did not contain JSON:\n{stdout}")
    return json.loads(stdout[json_start:])


def build_option_maps(describe_payload: dict[str, Any]) -> dict[str, dict[str, str]]:
    fields = describe_payload["data"]["describe"]["fields"]
    maps: dict[str, dict[str, str]] = {}
    for api_name in ("category", "product_line", "life_status", "product_status", "field_2wOsi__c"):
        field = fields.get(api_name) or {}
        options = field.get("options") or []
        maps[api_name] = {
            str(item.get("value")): str(item.get("label"))
            for item in options
            if item.get("value") is not None and item.get("label")
        }
    return maps


def to_iso(ms: Any) -> str | None:
    if ms in (None, ""):
        return None
    try:
        ts_ms = int(ms)
    except (TypeError, ValueError):
        return None
    return datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).isoformat()


def infer_fit(*parts: Any) -> str:
    haystack = " ".join(str(part or "") for part in parts).lower()
    if any(keyword in haystack for keyword in ENERGY_HIGH_KEYWORDS):
        return "HIGH"
    if any(keyword in haystack for keyword in ENERGY_MEDIUM_KEYWORDS):
        return "MEDIUM"
    return "LOW"


def load_mapping_pairs() -> list[dict[str, str]]:
    mapping_path = next((path for path in MAPPING_CANDIDATES if path.exists()), None)
    if mapping_path is None:
        return []
    workbook = load_workbook(mapping_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    pairs: list[dict[str, str]] = []
    rows = list(worksheet.iter_rows(values_only=True))
    for row in rows[1:]:
        left = "" if row[0] is None else str(row[0]).strip()
        right = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        if left and right:
            pairs.append({
                "fxiaokeField": left,
                "powerquoteField": right,
            })
    return pairs


def map_status(item: dict[str, Any]) -> tuple[str, str]:
    status_value = str(item.get("productStatusValue") or "").strip()
    if status_value == "1":
        return "ACTIVE", "上架"
    if status_value == "2":
        return "INACTIVE", "下架"
    if item.get("lifeStatusValue") == "normal" and item.get("isSaleable") is True:
        return "ACTIVE", "上架"
    return "INACTIVE", item.get("productStatusLabel") or "下架"


def build_directory_capability(item: dict[str, Any]) -> str:
    return str(item.get("categoryLabel") or item.get("categoryValue") or "")


def build_powerquote_models(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    models: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        status, status_label = map_status(item)
        models.append({
            "id": index,
            "crmId": item["crmId"],
            "modelCode": item.get("model") or "",
            "modelName": item["name"],
            "directoryCapability": build_directory_capability(item),
            "basePrice": float(item["price"]) if item.get("price") not in (None, "") else 0,
            "baseCost": None,
            "description": item.get("briefDescription") or item.get("description"),
            "status": status,
            "statusLabel": status_label,
            "prototypeFit": item["prototypeFit"],
            "sourceFields": {
                "model": item.get("model"),
                "productCode": item.get("productCode"),
                "masterCode": item.get("masterCode"),
                "categoryLabel": item.get("categoryLabel"),
                "categoryValue": item.get("categoryValue"),
                "productLineLabel": item.get("productLineLabel"),
                "productTypeLabel": item.get("productTypeLabel"),
                "specification": item.get("specification"),
                "price": item.get("price"),
                "lifeStatusLabel": item.get("lifeStatusLabel"),
                "isSaleable": item.get("isSaleable"),
            },
        })
    return models


def build_category_choices(
    describe_payload: dict[str, Any],
    items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    for item in items:
        value = str(item.get("categoryValue") or "").strip()
        if value:
            counts[value] = counts.get(value, 0) + 1

    field = ((describe_payload.get("data") or {}).get("describe") or {}).get("fields", {}).get("category", {})
    options = field.get("options") or []
    choices: list[dict[str, Any]] = []
    for option in options:
        value = str(option.get("value") or "").strip()
        label = str(option.get("label") or value).strip()
        if not value:
            continue
        choices.append({
            "value": value,
            "label": label,
            "count": counts.get(value, 0),
        })
    return choices


def main() -> None:
    describe = run_tool("describe", "--object", "产品")
    query = run_tool(
        "query",
        "--object",
        "产品",
        "--all-pages",
        "--page-size",
        "100",
        "--fields",
        FIELDS,
    )
    option_maps = build_option_maps(describe)
    rows = query["data"]["dataList"]
    mapping_pairs = load_mapping_pairs()

    items: list[dict[str, Any]] = []
    for row in rows:
        name = row.get("name")
        category_value = row.get("category")
        product_line_value = row.get("product_line")
        life_status_value = row.get("life_status")
        product_status_value = row.get("product_status")
        product_type_value = row.get("field_2wOsi__c")
        category_label = MANUAL_CATEGORY_LABELS.get(
            str(category_value),
            option_maps["category"].get(str(category_value), None if category_value is None else str(category_value)),
        )
        product_line_label = option_maps["product_line"].get(str(product_line_value), None if product_line_value is None else str(product_line_value))
        life_status_label = option_maps["life_status"].get(str(life_status_value), None if life_status_value is None else str(life_status_value))
        product_status_label = option_maps["product_status"].get(str(product_status_value), None if product_status_value is None else str(product_status_value))
        product_type_label = option_maps["field_2wOsi__c"].get(str(product_type_value), None if product_type_value is None else str(product_type_value))
        fit = infer_fit(
            name,
            category_label,
            product_line_label,
            product_type_label,
            row.get("product_spec"),
            row.get("field_uo1QB__c"),
        )
        items.append({
            "crmId": row.get("_id"),
            "name": name,
            "model": row.get("model"),
            "englishName": row.get("product_name_english__c"),
            "productCode": row.get("product_code"),
            "masterCode": row.get("field_93IC1__c"),
            "categoryValue": category_value,
            "categoryLabel": category_label,
            "productLineValue": product_line_value,
            "productLineLabel": product_line_label,
            "productTypeValue": product_type_value,
            "productTypeLabel": product_type_label,
            "specification": row.get("product_spec"),
            "price": row.get("price"),
            "isSaleable": row.get("is_saleable"),
            "lifeStatusValue": life_status_value,
            "lifeStatusLabel": life_status_label,
            "productStatusValue": product_status_value,
            "productStatusLabel": product_status_label,
            "onShelvesTime": to_iso(row.get("on_shelves_time")),
            "briefDescription": row.get("field_uo1QB__c"),
            "description": row.get("description"),
            "imageUrl": row.get("picture_path"),
            "prototypeFit": fit,
            "source": "Fxiaoke ProductObj",
        })

    mapped_models = build_powerquote_models(items)

    payload = {
        "syncedAt": datetime.now(timezone.utc).isoformat(),
        "sourceObject": "ProductObj",
        "total": len(items),
        "highFitCount": sum(1 for item in items if item["prototypeFit"] == "HIGH"),
        "mediumFitCount": sum(1 for item in items if item["prototypeFit"] == "MEDIUM"),
        "mapping": mapping_pairs,
        "categoryChoices": build_category_choices(describe, items),
        "items": items,
        "mappedModels": mapped_models,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"[OK] Exported {len(items)} products to {OUTPUT}")


if __name__ == "__main__":
    main()
